import os
import asyncio
import aiohttp
from bs4 import BeautifulSoup
import openpyxl
import socket

def is_connected(host="8.8.8.8", port=53, timeout=3):
    """
    Checks for network connectivity by attempting to connect to a well-known DNS server.
    Returns True if connected, else False.
    """
    try:
        socket.setdefaulttimeout(timeout)
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.connect((host, port))
        s.close()
        return True
    except Exception:
        return False

def create_workbook(workbook_path):
    """
    Creates a new Excel workbook (.xlsx) using openpyxl and saves it.
    """
    wb = openpyxl.Workbook()
    wb.save(os.path.abspath(workbook_path))

class TotoDataFetcher:
    def __init__(self, workbook_path="database_analysis.xlsx"):
        """
        Initializes the TotoDataFetcher with the given workbook path.
        """
        self.workbook_path = os.path.abspath(workbook_path)
        self.wb = None
        self.ws = None

    def load_or_create_workbook(self):
        """
        Loads the existing workbook if it exists; otherwise, creates a new Excel workbook.
        """
        if not os.path.exists(self.workbook_path):
            print(f"Workbook '{self.workbook_path}' not found. Creating a new workbook...")
            create_workbook(self.workbook_path)
        self.wb = openpyxl.load_workbook(self.workbook_path)
        # Use the first worksheet and rename it to "Data"
        self.ws = self.wb.active
        self.ws.title = "Data"

    def _clear_sheet(self):
        """
        Clears all cells in the worksheet.
        """
        for row in self.ws.iter_rows():
            for cell in row:
                cell.value = None

    def _fill_header_row(self):
        """
        Writes the header row to the sheet.
        """
        headers = ["Draw no.", "Date", "no1", "no2", "no3", "no4", "no5", "no6", "Addict. no"]
        self.ws.append(headers)

    def read_current_data(self):
        """
        Reads current data (excluding the header row) from the workbook and returns a list of tuples.
        """
        data = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                data.append(tuple(row))
        return data

    async def fetch_page(self, session, page):
        """
        Asynchronously fetches the HTML content of a given page.
        """
        url = f"https://en.lottolyzer.com/history/singapore/toto/page/{page}/per-page/50/summary-view"
        print(f"Fetching page {page} from {url}")
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/98.0.4758.102 Safari/537.36"
            )
        }
        try:
            async with session.get(url, headers=headers, timeout=10) as resp:
                if resp.status != 200:
                    print(f"Page {page} returned status {resp.status}")
                    return page, None
                text = await resp.text()
                print(f"Page {page} fetched with {len(text)} characters.")
                return page, text
        except Exception as e:
            print(f"Error fetching page {page}: {e}")
            return page, None

    async def get_all_draws(self, max_pages=50):
        """
        Asynchronously fetches draw data from pages 1..max_pages.
        Returns the list of draws sorted in descending order by 'draw'.
        """
        all_draws = {}
        async with aiohttp.ClientSession() as session:
            tasks = [self.fetch_page(session, page) for page in range(1, max_pages + 1)]
            pages = await asyncio.gather(*tasks)
            for page, text in pages:
                if text is None:
                    continue
                soup = BeautifulSoup(text, "html.parser")
                rows = soup.find_all("tr")
                page_new_draws = 0
                for row in rows:
                    cells = row.find_all("td")
                    if len(cells) < 4:
                        continue
                    try:
                        draw_num = int(cells[0].get_text(strip=True))
                    except ValueError:
                        continue
                    date_val = cells[1].get_text(strip=True)
                    win_val = cells[2].get_text(strip=True)  # e.g. "3, 9, 10, 12, 27, 41, 50"
                    addl_val = cells[3].get_text(strip=True)  # e.g. "29"
                    winning_numbers = [num.strip() for num in win_val.split(",") if num.strip()]
                    if draw_num not in all_draws:
                        all_draws[draw_num] = {
                            "draw": draw_num,
                            "date": date_val,
                            "winning_numbers": winning_numbers,
                            "addl": addl_val
                        }
                        page_new_draws += 1
                print(f"Page {page}: Extracted {page_new_draws} draws.")
                if page_new_draws == 0:
                    print(f"No new draws found on page {page}.")
                    break
        sorted_draws = sorted(all_draws.values(), key=lambda x: x["draw"], reverse=True)
        return sorted_draws

    def fill_entire_data_sheet(self, all_draws):
        """
        Fills the sheet with draw data (using the first six winning numbers).
        """
        for draw_info in all_draws:
            winning_numbers = draw_info["winning_numbers"][:6]
            row_data = (draw_info["draw"], draw_info["date"]) + tuple(winning_numbers) + (draw_info["addl"],)
            self.ws.append(row_data)

    def get_website_data_as_tuples(self, all_draws):
        """
        Converts the website draw data into a list of tuples for comparison.
        """
        data_tuples = []
        for draw_info in all_draws:
            winning_numbers = draw_info["winning_numbers"][:6]
            row_tuple = (draw_info["draw"], draw_info["date"]) + tuple(winning_numbers) + (draw_info["addl"],)
            data_tuples.append(row_tuple)
        return data_tuples

    def run(self):
        """
        Main method: checks for network connectivity, loads the workbook, fetches website data asynchronously,
        compares it with the current data, and if the data differs, deletes the current file and creates a new workbook
        with the latest data.
        """
        if not is_connected():
            print("Please on your network to get latest data for analysis.")
            return

        self.load_or_create_workbook()
        current_data = self.read_current_data()
        print("Current data in workbook:", current_data)
        print("Fetching all draws asynchronously from the website...")
        all_draws = asyncio.run(self.get_all_draws(max_pages=50))
        website_data = self.get_website_data_as_tuples(all_draws)
        print("Website data:", website_data)

        if current_data == website_data:
            print("Data is up-to-date. No changes made.")
        else:
            print("Data differs. Deleting current file and updating workbook with the latest data...")
            # Delete current file
            if os.path.exists(self.workbook_path):
                try:
                    os.remove(self.workbook_path)
                    print(f"Deleted existing file: {self.workbook_path}")
                except PermissionError as e:
                    print(f"PermissionError: {e}. Please close the file and try again.")
                    return
            # Create new workbook and update with website data
            create_workbook(self.workbook_path)
            self.load_or_create_workbook()
            self._fill_header_row()
            self.fill_entire_data_sheet(all_draws)
            self.wb.save(self.workbook_path)
            print(f"Workbook updated: {self.workbook_path}")

if __name__ == "__main__":
    analyzer = TotoDataFetcher()
    analyzer.run()
